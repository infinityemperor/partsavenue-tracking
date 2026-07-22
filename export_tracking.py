import os
import re
import sys
import json
import shutil
import hashlib
import subprocess
from datetime import datetime
import openpyxl

SVOD_PATH       = r'C:\Users\zinov\OneDrive\Desktop\РАБОЧАЯ\001 ИМПОРТ\006 excel учет\СВОД _new.xlsm'

# Путь к своду можно передать первым аргументом — так кнопка «Обновить трекинг»
# из книги гонит данные именно из той книги, откуда её нажали.
# Без аргумента поведение прежнее: берётся SVOD_PATH выше.
if len(sys.argv) > 1 and sys.argv[1].strip():
    SVOD_PATH = sys.argv[1].strip().strip('"')
    if not os.path.exists(SVOD_PATH):
        sys.exit(f'Не найден файл свода: {SVOD_PATH}')
print(f'Свод: {SVOD_PATH}')
CONTAINERS_ROOT = r'C:\Users\zinov\OneDrive\Desktop\РАБОЧАЯ\001 ИМПОРТ\001 containers\2026\001 UAE'
REPO_PATH       = r'C:\Users\zinov\OneDrive\Desktop\РАБОЧАЯ\001 ИМПОРТ\007 container tracing'
FILES_DIR       = os.path.join(REPO_PATH, 'files')


def read_tracking():
    wb = openpyxl.load_workbook(SVOD_PATH, data_only=True)
    ws = wb['Трекинг']
    result = []
    for row in ws.iter_rows(min_row=5, values_only=True):
        if not row[0] or not str(row[0]).startswith('UAE'):
            continue
        result.append({
            'name':            str(row[0]).strip(),
            'ship_plan':       row[1],
            'ship_fact':       row[2],
            'arrival_plan':    row[3],
            'arrival_fact':    row[4],
            'customs_plan':    row[5],
            'customs_fact':    row[6],
            'warehouse_plan':  row[7],
            'warehouse_fact':  row[8],
            'status':          str(row[9]) if row[9] else '',
        })
    return result


def find_container_folder(name):
    for folder in os.listdir(CONTAINERS_ROOT):
        if 'DO_NOT_COUNT' in folder.upper():
            continue
        code = folder.split('_')[0]
        if code == name:
            return os.path.join(CONTAINERS_ROOT, folder)
    return None


def find_latest_fpz(folder):
    files = [f for f in os.listdir(folder)
             if 'ФПЗ' in f and not f.startswith('~$') and f.lower().endswith('.xlsx')]
    if not files:
        return None

    def sort_key(f):
        m = re.search(r'(\d{2})\.(\d{2})\.(\d{4})', f)
        if m:
            d, mo, y = m.groups()
            return datetime(int(y), int(mo), int(d))
        return datetime.fromtimestamp(os.path.getmtime(os.path.join(folder, f)))

    files.sort(key=sort_key, reverse=True)
    return os.path.join(folder, files[0])


def find_ffz(folder):
    files = [f for f in os.listdir(folder)
             if 'ФФЗ' in f and not f.startswith('~$') and f.lower().endswith('.xlsx')]
    return os.path.join(folder, files[0]) if files else None


def file_hash(path):
    h = hashlib.md5()
    with open(path, 'rb') as f:
        h.update(f.read())
    return h.hexdigest()


def smart_copy(src, dest_dir, container_prefix, file_type):
    """
    Копирует файл в dest_dir только если изменился.
    Удаляет старую версию если имя файла изменилось.
    Возвращает относительный URL или None.
    """
    src_name = os.path.basename(src)
    dest_path = os.path.join(dest_dir, src_name)

    # Найти старые файлы этого типа для этого контейнера
    pattern = re.compile(rf'^{re.escape(container_prefix)}.*{file_type}.*\.xlsx$', re.IGNORECASE)
    old_files = [f for f in os.listdir(dest_dir) if pattern.match(f) and f != src_name]

    if os.path.exists(dest_path):
        # Файл с таким именем уже есть — проверяем хеш
        if file_hash(src) == file_hash(dest_path):
            print(f'  {src_name} — без изменений, пропускаем')
            return 'files/' + src_name
        else:
            print(f'  {src_name} — обновился, перезаписываем')
    else:
        # Новое имя файла — удаляем старые версии
        for old in old_files:
            os.remove(os.path.join(dest_dir, old))
            print(f'  Удалён старый файл: {old}')
        print(f'  Копируем: {src_name}')

    shutil.copy2(src, dest_path)
    return 'files/' + src_name


def read_fpz_items(fpz_path):
    """Читает состав из ФПЗ: Артикул, Бренд, Наименование, Количество."""
    try:
        wb = openpyxl.load_workbook(fpz_path, data_only=True, read_only=True)
        ws = wb.active
        items = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                break
            art  = str(row[0]).strip() if row[0] else ''
            brand = str(row[1]).strip() if row[1] else ''
            name  = str(row[2]).strip() if row[2] else ''
            qty   = row[4] if len(row) > 4 else None
            if art:
                items.append({'art': art, 'brand': brand, 'name': name, 'qty': qty})
        wb.close()
        return items
    except Exception as e:
        print(f'  Ошибка чтения состава ФПЗ: {e}')
        return []


def fmt(d):
    if d is None:
        return None
    if isinstance(d, datetime):
        return d.strftime('%Y-%m-%d')
    return str(d)


def calc_delays(c):
    today = datetime.today().replace(hour=0, minute=0, second=0, microsecond=0)
    delays = {}
    pairs = [
        ('ship',      c['ship_plan'],      c['ship_fact']),
        ('arrival',   c['arrival_plan'],   c['arrival_fact']),
        ('customs',   c['customs_plan'],   c['customs_fact']),
        ('warehouse', c['warehouse_plan'], c['warehouse_fact']),
    ]
    for key, plan, fact in pairs:
        if plan and not fact and isinstance(plan, datetime):
            delta = (today - plan).days
            if delta > 0:
                delays[key] = delta
    return delays


def main():
    os.makedirs(FILES_DIR, exist_ok=True)

    containers = read_tracking()
    output = []

    for c in containers:
        name = c['name']
        # Префикс для поиска файлов: UAE#5 -> UAE_005, UAE#6 -> UAE_006
        num = re.search(r'#(\d+)', name)
        file_prefix = f"UAE_{int(num.group(1)):03d}" if num else name.replace('#', '_')

        folder = find_container_folder(name)

        fpz_url = None
        ffz_url = None
        items   = []

        if folder:
            fpz = find_latest_fpz(folder)
            if fpz:
                fpz_url = smart_copy(fpz, FILES_DIR, file_prefix, 'ФПЗ')
                items = read_fpz_items(fpz)

            ffz = find_ffz(folder)
            if ffz:
                ffz_url = smart_copy(ffz, FILES_DIR, file_prefix, 'ФФЗ')
                if not items:  # нет ФПЗ — читаем состав из ФФЗ
                    items = read_fpz_items(ffz)

        output.append({
            'name':           name,
            'status':         c['status'],
            'ship_plan':      fmt(c['ship_plan']),
            'ship_fact':      fmt(c['ship_fact']),
            'arrival_plan':   fmt(c['arrival_plan']),
            'arrival_fact':   fmt(c['arrival_fact']),
            'customs_plan':   fmt(c['customs_plan']),
            'customs_fact':   fmt(c['customs_fact']),
            'warehouse_plan': fmt(c['warehouse_plan']),
            'warehouse_fact': fmt(c['warehouse_fact']),
            'delays':         calc_delays(c),
            'fpz_url':        fpz_url,
            'ffz_url':        ffz_url,
            'items':          items,
        })

    data = {
        'updated_at': datetime.now().strftime('%d.%m.%Y %H:%M'),
        'containers': output,
    }

    json_path = os.path.join(REPO_PATH, 'data.json')
    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    print(f'\ndata.json готов — {len(output)} контейнеров')

    subprocess.run(['git', '-C', REPO_PATH, 'add', '-A'], check=True)
    result = subprocess.run(
        ['git', '-C', REPO_PATH, 'commit', '-m',
         f'tracking update {datetime.now().strftime("%d.%m.%Y %H:%M")}'],
        capture_output=True, text=True
    )
    if 'nothing to commit' in result.stdout + result.stderr:
        print('Нет изменений для публикации')
    else:
        subprocess.run(['git', '-C', REPO_PATH, 'push'], check=True)
        print('Опубликовано на GitHub Pages')


if __name__ == '__main__':
    main()
